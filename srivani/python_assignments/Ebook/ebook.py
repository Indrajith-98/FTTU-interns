import os
import openpyxl

class UserBook:
    def __init__(self, filename="user_book_data.xlsx"):
        self.filename = filename
        self.users_books = self.load_data()

    def load_data(self):
        """Loads user and book data from the Excel file."""
        if not os.path.exists(self.filename):
            return {}
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook.active
        users_books = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            username, password, page, topic = row
            if username not in users_books:
                users_books[username] = {"password": password, "books": {}}
            book_title = f"Book {len(users_books[username]['books']) + 1}"  # Assign a default book title
            users_books[username]["books"][book_title] = {"page": page, "topic": topic}
        return users_books

    def save_data(self):
        """Saves user and book data to the Excel file."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Username", "Password", "Page", "Topic"])  # Adding headers
        
        for username, info in self.users_books.items():
            for book_title, book_info in info["books"].items():
                sheet.append([username, info["password"], book_info["page"], book_info["topic"]])
        
        workbook.save(self.filename)

    def register_user(self, username, password):
        """Registers a new user."""
        if username in self.users_books:
            print("Username already exists.")
        else:
            self.users_books[username] = {"password": password, "books": {}}
            self.save_data()
            print("User registered successfully.")

    def authenticate_user(self, username, password):
        """Authenticates a user."""
        if username in self.users_books and self.users_books[username]["password"] == password:
            return True
        return False

    def update_user(self, username, book_title, page, topic):
        """Updates the user's book info."""
        if username in self.users_books and book_title in self.users_books[username]["books"]:
            self.users_books[username]["books"][book_title]["page"] = page
            self.users_books[username]["books"][book_title]["topic"] = topic
            self.save_data()

    def add_book(self, username, book_title, page, topic):
        """Adds a new book to the user's collection."""
        if username in self.users_books:
            self.users_books[username]["books"][book_title] = {"page": page, "topic": topic}
            self.save_data()
            print(f"Book '{book_title}' added successfully!")
        else:
            print("User not found.")

    def get_user_info(self, username):
        """Retrieves information about the user."""
        if username in self.users_books:
            return self.users_books[username]
        else:
            return {"password": "", "books": {}}

    def display_user_info(self, username):
        """Displays user information for all books."""
        info = self.get_user_info(username)
        if info["books"]:
            for book_title, book_info in info["books"].items():
                print(f"Book: {book_title}")
                print(f"  Current Page: {book_info['page']}")
                print(f"  Current Topic: {book_info['topic']}")
        else:
            print(f"No books found for {username}.")

def user_login_or_register():
    user_book = UserBook()
    while True:
        action = input("Do you want to (l)ogin or (r)egister? (q to quit): ").lower()
        
        if action == 'q':
            break

        username = input("Enter username: ")
        password = input("Enter password: ")

        if action == 'l':
            if user_book.authenticate_user(username, password):
                print(f"Welcome back {username}!")
                return username  # Break out of the loop and return the username after login

            else:
                print("Invalid credentials. Please try again.")

        elif action == 'r':
            user_book.register_user(username, password)
            return username  # Break out of the loop and return the username after registration
        else:
            print("Invalid action. Please choose 'l' to login or 'r' to register.")

def main():
    username = user_login_or_register()
    user_book = UserBook()
    
    while True:
        action = input("Do you want to (r)ead, (u)pdate, (a)dd a book, or (q)uit?: ").lower()

        if action == 'q':
            break

        if action == 'r':
            user_book.display_user_info(username)
        elif action == 'u':
            print("Here are your books:")
            user_info = user_book.get_user_info(username)
            if not user_info["books"]:
                print("No books to update.")
                continue
            for idx, book_title in enumerate(user_info["books"], 1):
                print(f"{idx}. {book_title}")
            choice = int(input("Select a book to update (enter number): "))
            book_titles = list(user_info["books"].keys())
            if 1 <= choice <= len(book_titles):
                book_title = book_titles[choice - 1]
                page = int(input(f"Enter the page number you're on for '{book_title}': "))
                topic = input(f"Enter the current topic you're reading in '{book_title}': ")
                user_book.update_user(username, book_title, page, topic)
                print("Information updated!")
            else:
                print("Invalid choice.")
        elif action == 'a':
            book_title = input("Enter the title of the new book: ")
            page = int(input(f"Enter the page number for '{book_title}': "))
            topic = input(f"Enter the current topic you're reading in '{book_title}': ")
            user_book.add_book(username, book_title, page, topic)
        else:
            print("Invalid action. Please choose 'r' to read, 'u' to update, 'a' to add a book, or 'q' to quit.")

if __name__ == "__main__":
    main()
