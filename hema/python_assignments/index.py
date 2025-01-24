from openpyxl import Workbook, load_workbook
import os

#Create a book class
class EbookSystem:
    def __init__(self, books_file="books.xlsx", users_file="users.xlsx"):
        self.books_file = books_file #contains the details of the books available
        self.users_file = users_file #contains the details of the users and other related data
        self._initialize_books()
        self._initialize_users()

    # adding some sample data of available books to the excel sheets 
    def _initialize_books(self): 
        if not os.path.exists(self.books_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Books"
            ws.append(["Book Name", "Number of Pages", "Topics"])  # Add headers
            # Sample books data
            sample_books = [
                ["Python Basics", 3, "Introduction,Variables,Control Flow"],
                ["Advanced Python", 4, "OOP,Decorators,Generators,Async Programming"],
                ["Data Science", 5, "Intro,Data Wrangling,EDA,ML,DL"],
            ]
            for book in sample_books:
                ws.append(book)
            wb.save(self.books_file)

    #Adding the headers to the users excel sheet
    def _initialize_users(self):
        if not os.path.exists(self.users_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Users"
            ws.append(["User Name", "Book Name", "Number of Pages", "Current Page", "Topic"])  # Add headers
            wb.save(self.users_file)

    #adding a user and other details to user excel sheet if the book taken by the person is available 
    def add_user(self, user_name, book_name):
        # Add a new user and associate a book
        books_wb = load_workbook(self.books_file)
        books_ws = books_wb["Books"]

        # Find the selected book
        book_found = False
        for row in books_ws.iter_rows(min_row=2, values_only=True):
            if row[0] == book_name:
                num_pages = row[1]
                topics = row[2]
                book_found = True
                break
        books_wb.close()

        if not book_found:
            print(f"Book '{book_name}' not found!")
            return

        # Add user to the users sheet
        users_wb = load_workbook(self.users_file)
        users_ws = users_wb["Users"]

        for row in users_ws.iter_rows(min_row=2, values_only=True):
            if row[0] == user_name:
                print(f"User '{user_name}' already exists!")
                users_wb.close()
                return
        
        #add the details of the book related to that particular user
        users_ws.append([user_name, book_name, num_pages, 0, topics.split(",")[0]])
        users_wb.save(self.users_file)
        users_wb.close()
        print(f"User '{user_name}' added with the book '{book_name}'.")

    #update the details of the user if that person turns the page
    def turn_page(self, user_name):
    # Turn the page for a user
        users_wb = load_workbook(self.users_file)
        users_ws = users_wb["Users"]

    # Fetch book information from the books sheet
        books_wb = load_workbook(self.books_file)
        books_ws = books_wb["Books"]

        for row in users_ws.iter_rows(min_row=2):
            if row[0].value == user_name:
                current_page = row[3].value  # Current page
                book_name = row[1].value  # Book name
                num_pages = row[2].value  # Number of pages

            # Find the corresponding book in books.xlsx
                for book_row in books_ws.iter_rows(min_row=2, values_only=True):
                    if book_row[0] == book_name:  # Match book name
                        topics = book_row[2].split(",")  # Get topics list
                        break
                else:
                    print(f"Book '{book_name}' not found in books.xlsx!")
                    books_wb.close()
                    return

                if current_page + 1 < num_pages:  # Check if there's a next page
                    row[3].value = current_page + 1  # Increment the current page
                    topic = topics[current_page + 1]  # Fetch the next topic
                    row[4].value = topic  # Update the topic in users sheet
                    print(f"{user_name} turned to page {current_page + 2}: {topic}")
                else:
                    print(f"{user_name} has finished the book!")  # Last page reached
                break
        else:
            print(f"User '{user_name}' not found in users.xlsx!")

        users_wb.save(self.users_file)
        users_wb.close()
        books_wb.close()

    #printing the details of particular user
    def get_user_status(self, user_name):
        # Get the current status of a user
        users_wb = load_workbook(self.users_file)
        users_ws = users_wb["Users"]

        for row in users_ws.iter_rows(min_row=2, values_only=True):
            if row[0] == user_name:
                print(f"User: {user_name}\nBook: {row[1]}\nPages: {row[2]}\nCurrent Page: {row[3] + 1}\nTopic: {row[4]}")
                break
        else:
            print(f"User '{user_name}' not found!")

        users_wb.close()

# Example Usage
if __name__ == "__main__":
    system = EbookSystem()

    # Add users and assign books
    system.add_user("Alice", "Python Basics")
    system.add_user("Bob", "Advanced Python")

    # Get user status
    system.get_user_status("Alice")

    # Turn pages
    system.turn_page("Alice")
    system.get_user_status("Alice")
